from bs4 import BeautifulSoup
from requests_html import HTMLSession
import openpyxl

URL = "http://www.ebay.com/b/Home-Audio-Systems/184973/bn_115021122?rt=nc&_sop=12"

class Ebay_Seller():
    def init(self):
        self.url = str()
        self.location = str()
        self.date = str()
        self.positive = int()
        self.negative = int()
        self.neutral = int()

def get_html(url,params = None):
    session = HTMLSession()
    responce = session.get(url,params=params)
    if responce.ok:
        return responce.text

def get_url():
    for i in range(150,165):
        params = {"_pgn":i}
        a = set()
        html = get_html(URL,params=params)
        soup = BeautifulSoup(html,'lxml')
        li_tag = soup.find('ul',class_="b-list__items_nofooter srp-results srp-grid").find_all('li',class_="s-item s-item--large")
        for link_tag in li_tag:
            link_1 = link_tag.find('div',class_="s-item__info clearfix").find('a')['href']
            a.add(link_1)
    return a

def parse_1():
    linklar = get_url()
    b = set()
    for url_2 in linklar:
        html_1 = get_html(url_2)
        soup = BeautifulSoup(html_1,'lxml')
        try:
            nimadir = soup.find('div',attrs={"class":"ux-seller-section__item--seller","data-testid":"ux-seller-section__item--seller"}).find('a',attrs={"_sp":"p2047675.m3561.l2559","data-testid":"ux-action"})['href']
            b.add(nimadir)
        except:
            pass
    return b



def main():
    workbook = openpyxl.load_workbook('Ebay_3.xlsx')
    sheet = workbook['Seller_data']
    row_index = 42
    nimadir_2 = set()
    nimadir_2 = parse_1()
    Seller = Ebay_Seller()
    for url_1 in nimadir_2:
        html_2 = get_html(url_1)
        soup = BeautifulSoup(html_2,'lxml')
        location = soup.find('div',attrs={'id':"member_info","class":"mem_info"}).text
        o,p = location.split(': ')
        since,location_2 = p.split("|")
        if location_2 == "United States" and int(since[-4:])<2010:
            Seller.url = url_1
            Seller.location = location_2
            Seller.date = since
            positive_1 = soup.find('a',attrs={'_sp':'p2545226.m2531.l4717'}).find('div',class_='score').find('span',class_="num").text
            Seller.positive = positive_1
            neutral_1 = soup.find('a',attrs={'_sp':'p2545226.m2531.l4719'}).find('div',class_='score').find('span',class_='num').text
            Seller.neutral = neutral_1
            negative_1 = soup.find('a',attrs={'_sp':'p2545226.m2531.l4718'}).find('div',class_='score').find('span',class_='num').text
            Seller.negative = negative_1
            print(Seller.date,Seller.location,Seller.url,Seller.positive,Seller.neutral,Seller.negative)
            sheet.cell(row_index,2).value = Seller.date
            sheet.cell(row_index,3).value = Seller.location
            sheet.cell(row_index,4).value = Seller.url
            sheet.cell(row_index,5).value = Seller.positive
            sheet.cell(row_index,6).value = Seller.neutral
            sheet.cell(row_index,7).value = Seller.negative
            row_index += 1

    workbook.save('Ebay_3.xlsx')

if __name__ == "__main__":
    main()